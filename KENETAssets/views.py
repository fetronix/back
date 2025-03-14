from rest_framework.views import APIView
from rest_framework import viewsets,status, generics, permissions
from .models import *
from rest_framework.permissions import IsAuthenticated,AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import *  
from rest_framework.permissions import IsAdminUser
from django.views.generic import DetailView
from background_task import background
from rest_framework.generics import ListAPIView
from django.http import JsonResponse
from django.contrib.auth import logout
from django.contrib import messages
from rest_framework.response import Response
from rest_framework.decorators import api_view
from background_task.models import Task
from .tasks import process_rejected_cart_items, remove_expired_cart_items
from django.urls import reverse
from django.db import transaction
from django.utils import timezone
from django.http import HttpResponseNotFound,HttpResponse
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth import authenticate, login
import requests
from requests.auth import HTTPBasicAuth
from rest_framework.filters import SearchFilter
from rest_framework import viewsets, filters, status
from rest_framework.pagination import PageNumberPagination
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from django.core.files.base import ContentFile
import os
import base64
from django.conf import settings
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db.models import Q

from django.http import JsonResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.core.files.base import ContentFile
from .models import SavedPDF
import io
import openpyxl
from openpyxl.styles import Alignment
from django.http import HttpResponse
from .models import Location


class LoginView(APIView):
    permission_classes = [AllowAny]  # Allow any user to access this view

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        username = serializer.validated_data['username']
        password = serializer.validated_data['password']
        
        # Authenticate user
        user = authenticate(username=username, password=password)

        if user is not None:
            # Create JWT tokens for the user
            refresh = RefreshToken.for_user(user)
            return Response({
                "refresh": str(refresh),
                "access": str(refresh.access_token),
                "user": {
                    "id": user.id, 
                    "username": user.username,
                    "role": user.role,
                    "first_name": user.first_name,
                    "last_name": user.last_name,
                    "email": user.email,
                }
            }, status=status.HTTP_200_OK)
        else:
            return Response({"error": "Invalid credentials"}, status=status.HTTP_401_UNAUTHORIZED)

class AssetsPagination(PageNumberPagination):
    page_size = 10  # Set the number of assets per page
    page_size_query_param = 'page_size'  # Allow dynamic page size adjustment
    max_page_size = 100  # Set a max page size limit

class AssetsListView(ListAPIView):
    serializer_class = AssetsSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = AssetsPagination

    def get_queryset(self):
        # Filter queryset to include only items with status 'instore'
        return Assets.objects.filter(status='instore').order_by('-id')

class AssetsListViewAll(generics.ListAPIView):
    queryset = Assets.objects.all().order_by('-id')
    serializer_class = AssetsSerializer
    permission_classes = [IsAuthenticated]  # Only authenticated users can access
    # pagination_class = AssetsPagination

class DeliveryListView(generics.ListAPIView):
    queryset = Delivery.objects.all()
    serializer_class = DeliveryListSerializer
    permission_classes = [IsAuthenticated]  


@api_view(['POST'])
def create_or_update_location(request):
    location_name = request.data.get('location_name', None)
    location_alias = request.data.get('location_alias', None)

    if not location_name or not location_alias:
        return Response({"detail": "Location name and alias are required."}, status=status.HTTP_400_BAD_REQUEST)

    # Check if the location already exists
    location, created = Location.objects.get_or_create(
        name=location_name, 
        name_alias=location_alias
    )

    # Return appropriate response based on whether the location was created or updated
    if created:
        return Response(LocationSerializer(location).data, status=status.HTTP_201_CREATED)
    else:
        return Response(LocationSerializer(location).data, status=status.HTTP_200_OK)

class AssetsCreateView(generics.CreateAPIView):
    queryset = Assets.objects.all()
    serializer_class = AssetsSerializer
    permission_classes = [permissions.IsAuthenticated]

    def create(self, request, *args, **kwargs):
        # Handle location in assets creation
        location_data = request.data.get('location', None)

        if location_data:
            location_name = location_data.get('name', None)
            location_alias = location_data.get('name_alias', None)

            if location_name and location_alias:
                location, created = Location.objects.get_or_create(name=location_name, name_alias=location_alias)
                # Assign the created or existing location to the asset
                request.data['location'] = location.id

        return super().create(request, *args, **kwargs)

class AssetNewCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, *args, **kwargs):
        serializer = AssetCreateSerializer(data=request.data)
        if serializer.is_valid():
            asset = serializer.save()
            return Response(AssetsSerializer(asset).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class DeliveryNewCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, *args, **kwargs):
        serializer = DeliveryCreateSerializer(data=request.data)
        if serializer.is_valid():
            asset = serializer.save()
            return Response(DeliveryCreateSerializer(asset).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AssetsUpdateView(generics.UpdateAPIView):
    queryset = Assets.objects.all()
    serializer_class = AssetSerializer
    permission_classes = [permissions.IsAuthenticated]  # Ensure authentication

class AssetsDeleteView(generics.DestroyAPIView):
    queryset = Assets.objects.all()
    serializer_class = AssetsSerializer
    permission_classes = [permissions.IsAuthenticated]  # Ensure authentication
    
class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer # Ensure only authenticated users can access



class LocationViewSet(viewsets.ModelViewSet):
    queryset = Location.objects.all()
    serializer_class = LocationSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'name_alias']
    

class DeliveryViewSet(viewsets.ModelViewSet):
    queryset = Delivery.objects.all()
    serializer_class = DeliveryListSerializer # Ensure only authenticated users 
    
    
class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Suppliers.objects.all()
    serializer_class = SupplierSerializer # Ensure only authenticated users 

class UserViewSet(viewsets.ModelViewSet):
    queryset = CustomUser.objects.all()
    serializer_class = CustomUserSerializer

class UserCheckoutSet(viewsets.ModelViewSet):
    serializer_class = CustomUserCheckoutSerializer

    def get_queryset(self):
        """
        Exclude the logged-in user and optionally filter by role.
        """
        # Get the logged-in user
        logged_in_user = self.request.user

        # Get role parameter from query parameters
        role = self.request.query_params.get('role', None)

        # Base queryset: exclude the logged-in user
        queryset = CustomUser.objects.exclude(id=logged_in_user.id)

        if role:
            # Filter by role if provided
            return queryset.filter(role=role)

        # Default to returning users with 'can_checkout_items' role
        return queryset.filter(role=UserRoles.CAN_VERIFY_ITEMS)

    
class CartListView(generics.ListAPIView):
    serializer_class = CartSerializer
    

    def get_queryset(self):
        user = self.request.user
        
        # Optionally, you can call the task here if needed, e.g., for real-time expiration check
        remove_expired_cart_items(schedule=60)
        
        return Cart.objects.filter(user=user, asset__status='pending_release')
    

class AddToCartView(APIView):
    permission_classes = [IsAuthenticated]  # Ensure the user is authenticated

    def post(self, request, asset_id):
        user = request.user
        asset = get_object_or_404(Assets, id=asset_id)

        # Check if the asset is already in the user's cart
        if Cart.objects.filter(user=user, asset=asset).exists():
            return Response({'message': 'Item is already in the cart'}, status=status.HTTP_400_BAD_REQUEST)

        # Add to cart
        cart_item = Cart(user=user, asset=asset)
        cart_item.save()

        # Schedule the background task to remove expired cart items
        # remove_expired_cart_items()  # You can specify the delay as needed

        return Response({'message': 'Added to cart'}, status=status.HTTP_201_CREATED)



class RemoveFromCartView(APIView):
    permission_classes = [IsAuthenticated]  # Ensure the user is authenticated

    def delete(self, request, asset_id):
        user = request.user
        cart_item = get_object_or_404(Cart, id=asset_id)
        
        if cart_item:
            # Get the asset related to the cart item
            asset = get_object_or_404(Assets, id=cart_item.asset.id)  # Assuming Cart has a foreign key to Asset

            # Update the asset status to 'instore'
            asset.status = 'instore'
            asset.save()

            # Remove the item from the cart
            cart_item.delete()
            return Response({'message': 'Removed from Dispatch basket'}, status=status.HTTP_200_OK)
        else:
            return Response({'message': 'Item not in Dispatch basket'}, status=status.HTTP_400_BAD_REQUEST)



class CheckoutCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        user = request.user
        verified_user = request.data.get('verified_user')
        print("-----------------------------------------------")
        
        print(verified_user)
        
        print("-----------------------------------------------")
        # Filter cart items based on the status 'pending_release'
        cart_items = Cart.objects.filter(user=user, asset__status='pending_release')

        if not cart_items.exists():
            return Response({"detail": "No items with status 'pending_release' in your cart."}, status=status.HTTP_400_BAD_REQUEST)

        new_location = request.data.get('new_location')
        new_location1 = Location.objects.get(name=new_location)
        
        if not new_location:
            return Response({"detail": "New location is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Create a new Checkout for the current user
        checkout = Checkout.objects.create(user=user)
        checkout.cart_items.set(cart_items)
        
        v_user = CustomUser.objects.get(username=verified_user)
        
        checkout.verifier_user = CustomUser.objects.get(username=verified_user)
        
        checkout.save()

        # Update the status and location of the assets in the cart
        for cart_item in cart_items:
            asset = cart_item.asset
            asset.status = 'pending_approval'  # Update the status
            asset.destination_location = new_location1  # Set the new location
            asset.save()

        serializer = CheckoutSerializer(checkout)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class CheckoutListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated] 
    queryset = Checkout.objects.all()
    serializer_class = CheckoutSerializer

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)  # Show only the user's checkouts


class CheckoutPagination(PageNumberPagination):
    page_size = 10  # Set the number of assets per page
    page_size_query_param = 'page_size'  # Allow dynamic page size adjustment
    max_page_size = 100  # Set a max page size limit



class CheckoutAdminListView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = CheckoutSerializer
    # pagination_class = CheckoutPagination

    def get_queryset(self):
        # Get the logged-in user
        logged_in_user = self.request.user

        # Retrieve the first name and last name of the logged-in user
        user_first_name = logged_in_user.first_name
        user_last_name = logged_in_user.last_name
        username = logged_in_user.username
        usernameid = logged_in_user.id

        # Filter checkouts where the verifier_user matches the logged-in user's first or last name
        queryset = Checkout.objects.filter(
            models.Q(verifier_user=usernameid) | models.Q(verifier_user=usernameid)
        )

        if queryset.exists():
            return queryset
        else:
            # If no matching checkouts exist, return an empty queryset
            return Checkout.objects.none()


def get_base64_image(image_path):
    """
    Converts an image file to a base64 string.
    :param image_path: Path to the image file.
    :return: Base64-encoded string of the image.
    """
    full_path = os.path.join(settings.BASE_DIR, image_path)
    with open(full_path, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
    return f"data:image/png;base64,{encoded_string}"


class CheckoutDetailView(DetailView):
    model = Checkout
    template_name = 'kenet_release_form.html'  # Create this template for displaying the details
    context_object_name = 'checkout'

    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Convert images to base64 strings
        context['logo_base64'] = get_base64_image('static/assets/images/logo.png')
        context['stamp_base64'] = get_base64_image('static/assets/images/kenet_stamp.png')
        
        # Add user signature to the context
        checkout = self.get_object()  # Get the current Checkout object
        user_signature = checkout.user_signature_image
        admin_signature = checkout.signature_image

        # Check if the user signature exists and convert it to base64
        if user_signature:
            user_signature_base64 = get_base64_image(user_signature.path)
            context['user_signature_base64'] = user_signature_base64
        else:
            context['user_signature_base64'] = None  # or some default value

        # Check if the admin signature exists and convert it to base64
        if admin_signature:
            admin_signature_base64 = get_base64_image(admin_signature.path)
            context['admin_signature_base64'] = admin_signature_base64
        else:
            context['admin_signature_base64'] = None  # or some default value
                
        return context

    def get(self, request, *args, **kwargs):
        # Check for action parameter to save the PDF
        action = request.GET.get('action', None)
        if action == 'save_pdf':
            return self.generate_and_save_pdf()
        return super().get(request, *args, **kwargs)

    def generate_and_save_pdf(self):
        checkout = self.get_object()
        cart_items = checkout.cart_items.all()

        # Get base64 images for logo and stamp
        logo_base64 = get_base64_image('static/assets/images/logo.png')
        stamp_base64 = get_base64_image('static/assets/images/kenet_stamp.png')

        # Get user and admin signatures, if available
        user_signature = checkout.user_signature_image
        admin_signature = checkout.signature_image

        # Convert user and admin signatures to base64
        if user_signature:
            user_signature_base64 = get_base64_image(user_signature.path)
        else:
            user_signature_base64 = None  # or a default signature if needed

        if admin_signature:
            admin_signature_base64 = get_base64_image(admin_signature.path)
        else:
            admin_signature_base64 = None  # or a default signature if needed

        # Render the HTML template with all the necessary context (checkout, cart items, images)
        html_string = render_to_string(self.template_name, {
            'checkout': checkout,
            'cart_items': cart_items,
            'logo_base64': logo_base64,
            'stamp_base64': stamp_base64,
            'user_signature_base64': user_signature_base64,
            'admin_signature_base64': admin_signature_base64,
        })

        # Generate the PDF from the HTML string
        pdf = HTML(string=html_string).write_pdf()

        # Save the generated PDF to the model
        pdf_file = ContentFile(pdf)
        pdf_filename = f"release_form_{checkout.id}.pdf"
        checkout.pdf_file.save(pdf_filename, pdf_file, save=True)

        # Return a response to confirm the save
        return HttpResponse(f"PDF saved as {pdf_filename}.")


    
class ApproveCheckoutView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, checkout_id, *args, **kwargs):
        try:
            checkout = Checkout.objects.get(id=checkout_id)
            cart_items = checkout.cart_items.all()

            failed_assets = []
            for cart_item in cart_items:
                asset = cart_item.asset
                if asset.status == 'approved':
                    failed_assets.append(f"Asset {asset.serial_number} is already approved.")
                elif asset.status == 'pending_approval':
                    asset.status = 'approved'
                    asset.save()
                else:
                    failed_assets.append(f"Asset {asset.serial_number} cannot be approved due to its current status: {asset.status}.")

            if not failed_assets:
                # Generate the checkout detail URL link
                checkout_url = request.build_absolute_uri(reverse('checkout-detail', args=[checkout_id]))

                # Save the generated URL in the `checkout_url_link` field
                checkout.checkout_url_link = checkout_url
                checkout.save()

                # save_approved_asset_movements()

                return Response(
                    {
                        "detail": "All assets in the checkout have been approved.",
                        "checkout_url_link": checkout_url
                    },
                    status=status.HTTP_200_OK
                )

            return Response(
                {"detail": "Some assets could not be approved.", "failed_assets": failed_assets},
                status=status.HTTP_400_BAD_REQUEST
            )

        except Checkout.DoesNotExist:
            return Response({"detail": "Checkout not found."}, status=status.HTTP_404_NOT_FOUND)

class RejectCheckoutView(APIView):
    permission_classes = [IsAdminUser]  # Only admins can approve checkouts

    def post(self, request, checkout_id, *args, **kwargs):
        try:
            checkout = Checkout.objects.get(id=checkout_id)
            cart_items = checkout.cart_items.all()

            # List to keep track of assets that couldn't be approved
            failed_assets = []

            for cart_item in cart_items:
                asset = cart_item.asset
                if asset.status == 'rejected':
                    # If the asset is already rejected, add to the failed_assets list
                    failed_assets.append(f"Asset {asset.serial_number} is already rejected.")
                elif asset.status == 'pending_approval':
                    # Reject the asset if it's pending approval
                    asset.status = 'rejected'
                    asset.save()
                else:
                    # If asset status is not "pending_approval", you can either skip or handle it differently
                    failed_assets.append(f"Asset {asset.serial_number} cannot be rejected due to its current status: {asset.status}.")

            # If there are no failed assets, trigger the background task
            if not failed_assets:
                # Trigger the background task to handle cart and asset cleanup after rejection
                # Correct call to background task
                process_rejected_cart_items() 
                
                return JsonResponse(
                    {"detail": "All assets in the checkout have been rejected and processed."},
                    status=status.HTTP_200_OK
                )

            # If there are failed assets, return a message with all the failed ones
            return JsonResponse(
                {"detail": "Some assets could not be rejected.", "failed_assets": failed_assets},
                status=status.HTTP_400_BAD_REQUEST
            )

        except Checkout.DoesNotExist:
            return JsonResponse({"detail": "Checkout not found."}, status=status.HTTP_404_NOT_FOUND)



class CheckoutUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Checkout.objects.all()
    serializer_class = CheckoutUpdateSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_update(self, serializer):
        # Save the updated checkout instance
        checkout = serializer.save()
        
        # Iterate through each cart item linked to this checkout
        for cart_item in checkout.cart_items.all():
            asset = cart_item.asset
            
            # Create an AssetsMovement entry for each asset in the checkout cart
            AssetsMovement.objects.create(
                assets=asset,
                date_created=timezone.now(),
                person_moving=self.request.user,
                comments=checkout.remarks,
                serial_number=asset.serial_number,
                asset_description = asset.asset_description,
                asset_description_model = asset.asset_description_model,
                kenet_tag=asset.kenet_tag,
                status="onsite",  # Set the movement record status to "onsite"
                location=asset.location.name if asset.location else None,
                new_location=asset.destination_location,
            )

            # Update the asset status to "onsite"
            asset.status = "onsite"
            asset.save()

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        
        # Customize the response if needed
        response.data['message'] = "Checkout updated, asset status set to onsite, and movements recorded"
        return Response(response.data, status=status.HTTP_200_OK)

class CheckoutUSerUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Checkout.objects.all()
    serializer_class = CheckoutUserUpdateSerializer
    permission_classes = [permissions.IsAuthenticated]

   
    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        
        # Customize the response if needed
        response.data['message'] = "Checkout updated, asset status set to onsite, and movements recorded"
        return Response(response.data, status=status.HTTP_200_OK)


def custom_404(request, exception=None):
    return render(request, 'qyfy/404.html', status=404)

def custom_505(request, exception=None):
    return render(request, 'qyfy/505.html', status=404)


def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.first_name}!')
            return redirect('home')  # Redirect to the home page after login
        else:
            error_message = "Invalid username or password"
            return render(request, 'login.html', {'error_message': error_message})
    
    return render(request, 'login.html')

from django.db.models import Q, F
from django.db.models.functions import Lower

def home_view(request):
    search_query = request.GET.get('search', '').lower()  # Convert search query to lowercase
    status_filter = request.GET.get('status', '').lower()  # Convert status filter to lowercase
    location_filter = request.GET.get('location', '').lower()  # Convert location filter to lowercase

    # Start with all assets
    assets_list = Assets.objects.all()

    # Apply search query
    if search_query:
        assets_list = assets_list.filter(
            Q(asset_description__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(kenet_tag__icontains=search_query)
        )
    
    # Apply status filter
    if status_filter:
        assets_list = assets_list.filter(status__iexact=status_filter)

    # Apply location filter
    if location_filter:
        assets_list = assets_list.filter(
            Q(location__name__icontains=location_filter) | 
            Q(location__name_alias__icontains=location_filter) |
            Q(location__location_code__icontains=location_filter)
        )

    # Paginate with 10 assets per page
    paginator = Paginator(assets_list, 10)
    page_number = request.GET.get('page')  # Get current page
    assets = paginator.get_page(page_number)  # Get paginated assets

    return render(request, 'home.html', {
        'user': request.user,
        'assets': assets,
        'search_query': search_query,
        'status_filter': status_filter,
        'location_filter': location_filter,
    })
    
    
def view_all_pdfs(request):
    # Filter Checkouts with saved PDF files
    checkouts_with_pdfs = Checkout.objects.filter(pdf_file__isnull=False)
    
    context = {
        'checkouts': checkouts_with_pdfs
    }
    return render(request, 'view_pdfs.html', context)

def logout_view(request):
    logout(request)
    messages.info(request, "You have been successfully logged out.")
    return redirect('login-form')  # Redirect to the login page or home page


import logging

logger = logging.getLogger(__name__)


def kenet_release_form_view(request):
    # Example checkout object, replace with your logic to get the correct checkout
    checkout = get_object_or_404(Checkout, id=request.GET.get("id")) 

    # Initialize the variables in case the user is not found
    authorizing_first_name = "Unknown"
    authorizing_last_name = "User"

    authorizing_user = CustomUser.objects.get(username=checkout.authorizing_name)
    authorizing_first_name = authorizing_user.first_name
    authorizing_last_name = authorizing_user.last_name

    # Add to context
    context = {
        'checkout': checkout,
        'logo_url': '/static/assets/images/logo.png',
        'stamp_url': '/static/assets/images/kenet_stamp.png',
        'authorizing_first_name': authorizing_first_name,
        'authorizing_last_name': authorizing_last_name,
    }
    return render(request, 'kenet_release_form.html', context)

    

class ReturnFaultyAssetView(APIView):
    allowed_methods = ['GET', 'POST', 'PATCH']  # Allow PATCH as well
    
    def get(self, request, asset_id, *args, **kwargs):
        # Fetch and return asset details for GET requests
        asset = get_object_or_404(Assets, id=asset_id)
        # Assuming you have a serializer to return the asset data
        # serializer = AssetSerializer(asset)
        return Response({"asset_id": asset.id, "status": asset.status}, status=status.HTTP_200_OK)
    
    def post(self, request, asset_id, *args, **kwargs):
        # Mark the asset as faulty and remove it from cart/checkout
        return self.handle_faulty_asset(request, asset_id)
    
    def patch(self, request, asset_id, *args, **kwargs):
        # Handle partial updates (you can mark as faulty and remove from cart)
        return self.handle_faulty_asset(request, asset_id)
    
    def handle_faulty_asset(self, request, asset_id):
        try:
            # Get the asset by ID
            asset = get_object_or_404(Assets, id=asset_id)

            # Check if the asset's status is already 'faulty'
            if asset.status == 'faulty':
                return Response(
                    {"detail": "Asset is already marked as faulty."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Mark asset status as 'faulty'
            asset.status = 'faulty'
            
            AssetsMovement.objects.create(
                assets=asset,
                date_created=timezone.now(),
                person_moving=self.request.user,
                # comments=checkout.remarks,
                serial_number=asset.serial_number,
                asset_description = asset.asset_description,
                asset_description_model = asset.asset_description_model,
                kenet_tag=asset.kenet_tag,
                status=asset.status,  # Set the movement record status to "onsite"
                location=asset.location.name if asset.location else None,
                new_location="UON Store"
            )

            # Check if the asset exists in any checkout or cart
            cart_item = Cart.objects.filter(asset=asset).first()
            checkout = Checkout.objects.filter(cart_items__asset=asset).first()

            with transaction.atomic():
                if checkout:
                    # Remove the asset from the checkout
                    checkout.cart_items.remove(cart_item)
                    checkout.save()

                if cart_item:
                    # Delete the cart item (if not deleted above)
                    cart_item.delete()

                # Revert `new_location` to null
                asset.new_location = "UON Store"
                asset.save()

            return Response(
                {"detail": "Asset has been marked as faulty and removed from any checkouts or carts."},
                status=status.HTTP_200_OK
            )

        except Assets.DoesNotExist:
            return Response({"detail": "Asset not found."}, status=status.HTTP_404_NOT_FOUND)


class ReturnStoreAssetView(APIView):
    allowed_methods = ['GET', 'POST', 'PATCH']  # Allow PATCH as well
    
    def get(self, request, asset_id, *args, **kwargs):
        # Fetch and return asset details for GET requests
        asset = get_object_or_404(Assets, id=asset_id)
        # Assuming you have a serializer to return the asset data
        # serializer = AssetSerializer(asset)
        return Response({"asset_id": asset.id, "status": asset.status}, status=status.HTTP_200_OK)
    
    def post(self, request, asset_id, *args, **kwargs):
        # Mark the asset as faulty and remove it from cart/checkout
        return self.handle_faulty_asset(request, asset_id)
    
    def patch(self, request, asset_id, *args, **kwargs):
        # Handle partial updates (you can mark as faulty and remove from cart)
        return self.handle_faulty_asset(request, asset_id)
    
    def handle_faulty_asset(self, request, asset_id):
        try:
            # Get the asset by ID
            asset = get_object_or_404(Assets, id=asset_id)

            # Check if the asset's status is already 'faulty'
            if asset.status == 'instore':
                return Response(
                    {"detail": "Asset is already marked as instore."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Mark asset status as 'faulty'
            asset.status = 'instore'
            
            AssetsMovement.objects.create(
                assets=asset,
                date_created=timezone.now(),
                person_moving=self.request.user,
                # comments=checkout.remarks,
                serial_number=asset.serial_number,
                asset_description = asset.asset_description,
                asset_description_model = asset.asset_description_model,
                kenet_tag=asset.kenet_tag,
                status=asset.status,  # Set the movement record status to "onsite"
                location=asset.location.name if asset.location else None,
                new_location="UON Store"
            )

            # Check if the asset exists in any checkout or cart
            cart_item = Cart.objects.filter(asset=asset).first()
            checkout = Checkout.objects.filter(cart_items__asset=asset).first()

            with transaction.atomic():
                if checkout:
                    # Remove the asset from the checkout
                    checkout.cart_items.remove(cart_item)
                    checkout.save()

                if cart_item:
                    # Delete the cart item (if not deleted above)
                    cart_item.delete()

                # Revert `new_location` to null
                asset.new_location = "UON Store"
                asset.save()

            return Response(
                {"detail": "Asset has been marked as instore and removed from any checkouts or carts."},
                status=status.HTTP_200_OK
            )

        except Assets.DoesNotExist:
            return Response({"detail": "Asset not found."}, status=status.HTTP_404_NOT_FOUND)


class ReturnDecomissionedAssetView(APIView):
    allowed_methods = ['GET', 'POST', 'PATCH']  # Allow PATCH as well
    
    def get(self, request, asset_id, *args, **kwargs):
        # Fetch and return asset details for GET requests
        asset = get_object_or_404(Assets, id=asset_id)
        # Assuming you have a serializer to return the asset data
        # serializer = AssetSerializer(asset)
        return Response({"asset_id": asset.id, "status": asset.status}, status=status.HTTP_200_OK)
    
    def post(self, request, asset_id, *args, **kwargs):
        # Mark the asset as faulty and remove it from cart/checkout
        return self.handle_faulty_asset(request, asset_id)
    
    def patch(self, request, asset_id, *args, **kwargs):
        # Handle partial updates (you can mark as faulty and remove from cart)
        return self.handle_faulty_asset(request, asset_id)
    
    def handle_faulty_asset(self, request, asset_id):
        try:
            # Get the asset by ID
            asset = get_object_or_404(Assets, id=asset_id)

            # Check if the asset's status is already 'faulty'
            if asset.status == 'decommissioned':
                return Response(
                    {"detail": "Asset is already marked as decommissioned."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Mark asset status as 'faulty'
            asset.status = 'decommissioned'
            
            AssetsMovement.objects.create(
                assets=asset,
                date_created=timezone.now(),
                person_moving=self.request.user,
                # comments=checkout.remarks,
                serial_number=asset.serial_number,
                asset_description = asset.asset_description,
                asset_description_model = asset.asset_description_model,
                kenet_tag=asset.kenet_tag,
                status=asset.status,  # Set the movement record status to "onsite"
                location=asset.location.name if asset.location else None,
                new_location="UON Store"
            )

            # Check if the asset exists in any checkout or cart
            cart_item = Cart.objects.filter(asset=asset).first()
            checkout = Checkout.objects.filter(cart_items__asset=asset).first()

            with transaction.atomic():
                if checkout:
                    # Remove the asset from the checkout
                    checkout.cart_items.remove(cart_item)
                    checkout.save()

                if cart_item:
                    # Delete the cart item (if not deleted above)
                    cart_item.delete()

                # Revert `new_location` to null
                asset.destination_location = None
                asset.save()

            return Response(
                {"detail": "Asset has been marked as decommissioned and removed from any checkouts or carts."},
                status=status.HTTP_200_OK
            )

        except Assets.DoesNotExist:
            return Response({"detail": "Asset not found."}, status=status.HTTP_404_NOT_FOUND)




def download_locations_as_excel(request):
    # Create a workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Locations'

    # Add header row
    headers = ['ID', 'Name', 'Name Alias']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add data rows
    locations = Location.objects.all()
    for row_num, location in enumerate(locations, start=2):
        sheet.cell(row=row_num, column=1).value = location.id
        sheet.cell(row=row_num, column=2).value = location.name
        sheet.cell(row=row_num, column=3).value = location.name_alias

    # Set response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=locations.xlsx'

    # Save workbook to response
    workbook.save(response)
    return response


def save_pdf(request):
    if request.method == 'POST':
        # Get user from the request
        # user = request.user
        
        # Generate the PDF content from the template
        html = render_to_string('template_name.html', {'checkout': request.POST.get('checkout_data')})
        pdf_file = io.BytesIO()
        pisa_status = pisa.CreatePDF(html, dest=pdf_file)
        
        if not pisa_status.err:
            pdf_file.seek(0)
            pdf_filename = f"release_form.pdf"
            saved_pdf = SavedPDF.objects.create(
                # user=user,
                pdf_file=ContentFile(pdf_file.read(), name=pdf_filename)
            )
            return JsonResponse({'message': 'PDF saved successfully', 'pdf_id': saved_pdf.id}, status=201)
        else:
            return JsonResponse({'error': 'Failed to generate PDF'}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)



